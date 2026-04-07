#!/usr/bin/env python3
"""
Import MijnGS1 products from Excel into Interwall database.
Creates:
  1. Component products (CPUs, RAM sticks, SSDs, GPUs, Case, PSU)
  2. Composite products (every EAN from the spreadsheet)
  3. EAN compositions linking each composite to its components

Usage: docker exec -i interwall-postgres psql -U interwall -d interwall < import.sql
       (this script generates the SQL)
"""

import re
import sys
import openpyxl

XLSX_PATH = sys.argv[1] if len(sys.argv) > 1 else "/Users/ottogen/Downloads/8719965383890-products.xlsx"
SHEET_NAME = "10075055"

# =============================================================================
# Component catalog — these are the individual parts that go into builds
# =============================================================================
COMPONENTS = {
    # CPUs
    "CPU-R3-3200":    ("COMP-CPU-R3-3200",    "AMD Ryzen 3 3200G"),
    "CPU-R5-3400":    ("COMP-CPU-R5-3400",    "AMD Ryzen 5 3400G"),
    "CPU-R5-4500":    ("COMP-CPU-R5-4500",    "AMD Ryzen 5 4500"),
    "CPU-R7-5700":    ("COMP-CPU-R7-5700",    "AMD Ryzen 7 5700"),
    "CPU-R7-5700X":   ("COMP-CPU-R7-5700X",   "AMD Ryzen 7 5700X"),
    # RAM sticks (individual sticks — compositions use multiples)
    "RAM-8GB":        ("COMP-RAM-8GB",         "DDR4 8GB RAM Stick"),
    "RAM-16GB":       ("COMP-RAM-16GB",        "DDR4 16GB RAM Stick"),
    # SSDs
    "SSD-256GB":      ("COMP-SSD-256GB",       "NVMe SSD 256GB"),
    "SSD-500GB":      ("COMP-SSD-500GB",       "NVMe SSD 500GB"),
    "SSD-512GB":      ("COMP-SSD-512GB",       "NVMe SSD 512GB"),
    "SSD-1TB":        ("COMP-SSD-1TB",         "NVMe SSD 1TB"),
    "SSD-2TB":        ("COMP-SSD-2TB",         "NVMe SSD 2TB"),
    # GPUs
    "GPU-RTX3050":    ("COMP-GPU-RTX3050",     "GeForce RTX 3050"),
    "GPU-RTX3060":    ("COMP-GPU-RTX3060",     "GeForce RTX 3060 12GB"),
    "GPU-RTX4060":    ("COMP-GPU-RTX4060",     "GeForce RTX 4060"),
    "GPU-RTX5050":    ("COMP-GPU-RTX5050",     "GeForce RTX 5050"),
    "GPU-RTX5060":    ("COMP-GPU-RTX5060",     "GeForce RTX 5060"),
    "GPU-RTX5060TI":  ("COMP-GPU-RTX5060TI",   "GeForce RTX 5060 Ti"),
    "GPU-RTX5070":    ("COMP-GPU-RTX5070",     "GeForce RTX 5070"),
    "GPU-RTX5070T":   ("COMP-GPU-RTX5070T",    "GeForce RTX 5070 Ti"),
    # Fixed components (every build gets these)
    "CASE-NGG":       ("COMP-CASE-NGG",        "NGG Gaming Case"),
    "PSU-STD":        ("COMP-PSU-STD",         "Standard PSU 550W"),
    "MOBO-AM4":       ("COMP-MOBO-AM4",        "AM4 Motherboard"),
}

# RAM mapping: total GB → (stick EAN key, count)
RAM_MAP = {
    8:  ("RAM-8GB", 1),
    16: ("RAM-8GB", 2),
    32: ("RAM-16GB", 2),
    64: ("RAM-16GB", 4),
}

def parse_description(desc):
    """Parse a GS1 product description into component keys.
    Returns list of (component_key, quantity) or None if unparsable."""
    desc = desc.strip()
    components = []

    # Normalize
    d = desc.lower().replace("™", "").replace("®", "").strip()
    # Remove '-setup' suffix (same components, just includes Windows setup)
    d = re.sub(r'-setup$', '', d)

    # Skip the generic "Gaming PC" entry with no specs
    if d == "gaming pc":
        return None

    # ── CPU ──
    cpu = None
    if re.search(r'ryzen\s*7[- ]*5700x', d):
        cpu = "CPU-R7-5700X"
    elif re.search(r'ryzen\s*7[- ]*5700', d):
        cpu = "CPU-R7-5700"
    elif re.search(r'ryzen\s*5[- ]*4500', d):
        cpu = "CPU-R5-4500"
    elif re.search(r'ryzen\s*5[- ]*3400', d):
        cpu = "CPU-R5-3400"
    elif re.search(r'ryzen\s*3[- ]*3200', d):
        cpu = "CPU-R3-3200"

    if not cpu:
        return None  # Can't determine build without CPU
    components.append((cpu, 1))

    # ── RAM ──
    ram_gb = None
    m = re.search(r'(\d+)\s*gb\s*ram', d)
    if m:
        ram_gb = int(m.group(1))
    else:
        m = re.search(r'(\d+)gb', d)
        if m:
            ram_gb = int(m.group(1))

    if ram_gb and ram_gb in RAM_MAP:
        stick_key, count = RAM_MAP[ram_gb]
        components.append((stick_key, count))

    # ── SSD ──
    # Strategy: find the storage value (not the RAM value).
    # Long format: "1 TB SSD", "512 GB SSD"
    # Short format: "16gb-512gb", "16gb/1tb", "16gb-1000gb"
    # The RAM value is always smaller and already captured above.
    ssd = None
    m = re.search(r'(\d+)\s*tb\s*ssd', d)
    if m:
        ssd = f"SSD-{int(m.group(1))}TB"
    else:
        m = re.search(r'(\d+)\s*gb\s*ssd', d)
        if m:
            ssd = f"SSD-{int(m.group(1))}GB"
        else:
            # Short format: find all numbers with gb/tb suffix, pick the storage one
            # Storage indicators: after '/', after '-' following the RAM value, or >=100gb
            # Try /NTB or /NGB pattern first (e.g. 16GB/1TB)
            m = re.search(r'/(\d+)(gb|tb)', d)
            if m:
                val, unit = int(m.group(1)), m.group(2)
                if unit == 'tb':
                    ssd = f"SSD-{val}TB"
                else:
                    ssd = f"SSD-{val}GB" if val < 1000 else f"SSD-{val // 1000}TB"
            else:
                # Find all number-unit pairs, exclude the RAM one
                all_vals = re.findall(r'(\d+)(gb|tb|t)\b', d)
                for val_s, unit in all_vals:
                    val = int(val_s)
                    # Skip if this is the RAM value
                    if ram_gb and val == ram_gb:
                        continue
                    # Skip tiny values that aren't storage
                    if unit == 'gb' and val < 100:
                        continue
                    if unit in ('t', 'tb'):
                        ssd = f"SSD-{val}TB"
                    elif val >= 1000:
                        ssd = f"SSD-{val // 1000}TB"
                    else:
                        ssd = f"SSD-{val}GB"
                    break

    if ssd and ssd in COMPONENTS:
        components.append((ssd, 1))

    # ── GPU ──
    gpu = None
    if re.search(r'rtx\s*5070\s*t', d):
        gpu = "GPU-RTX5070T"
    elif re.search(r'rtx\s*5070', d):
        gpu = "GPU-RTX5070"
    elif re.search(r'rtx\s*5060\s*ti', d):
        gpu = "GPU-RTX5060TI"
    elif re.search(r'rtx\s*5060i', d):
        gpu = "GPU-RTX5060TI"  # rtx5060I = 5060 Ti
    elif re.search(r'rtx\s*5060', d):
        gpu = "GPU-RTX5060"
    elif re.search(r'rtx\s*5050', d):
        gpu = "GPU-RTX5050"
    elif re.search(r'rtx\s*4060', d):
        gpu = "GPU-RTX4060"
    elif re.search(r'rtx\s*3060', d):
        gpu = "GPU-RTX3060"
    elif re.search(r'rtx\s*3050', d):
        gpu = "GPU-RTX3050"
    # Integrated GPU (Vega) = no discrete GPU component

    if gpu:
        components.append((gpu, 1))

    # ── Fixed components (every build) ──
    components.append(("CASE-NGG", 1))
    components.append(("PSU-STD", 1))
    components.append(("MOBO-AM4", 1))

    return components


def sql_escape(s):
    """Escape single quotes for SQL."""
    return s.replace("'", "''")


def main():
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb[SHEET_NAME]

    print("-- =============================================================================")
    print("-- Auto-generated: MijnGS1 product import for Interwall")
    print("-- =============================================================================")
    print("BEGIN;")
    print()

    # 1. Create component products
    print("-- Component products (individual parts)")
    for key, (ean, name) in COMPONENTS.items():
        print(f"INSERT INTO products (ean, name, sku, is_composite) "
              f"VALUES ('{ean}', '{sql_escape(name)}', '{key}', FALSE) "
              f"ON CONFLICT (ean) DO NOTHING;")
    print()

    # 2. Create composite products from spreadsheet
    print("-- Composite products (assembled PCs from MijnGS1)")
    products = []
    skipped = 0
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        ean = str(row[0]).strip() if row[0] else None
        desc = str(row[6]).strip() if row[6] else None
        sub_brand = str(row[9]).strip() if row[9] else "NGG"
        if not ean or not desc:
            continue

        comps = parse_description(desc)
        if comps is None:
            skipped += 1
            continue

        is_setup = "-setup" in desc.lower() or sub_brand == "Setup"
        clean_name = re.sub(r'-setup$', '', desc, flags=re.IGNORECASE).strip()
        if is_setup:
            clean_name += " (Setup)"

        products.append((ean, clean_name, comps))
        print(f"INSERT INTO products (ean, name, sku, is_composite) "
              f"VALUES ('{ean}', '{sql_escape(clean_name)}', NULL, TRUE) "
              f"ON CONFLICT (ean) DO NOTHING;")

    print()
    print(f"-- {len(products)} composite products, {skipped} skipped (unparsable)")
    print()

    # 3. Create compositions
    print("-- EAN Compositions")
    comp_count = 0
    for ean, name, comps in products:
        for comp_key, qty in comps:
            comp_ean = COMPONENTS[comp_key][0]
            print(f"INSERT INTO ean_compositions (parent_ean, component_ean, quantity) "
                  f"VALUES ('{ean}', '{comp_ean}', {qty}) "
                  f"ON CONFLICT (parent_ean, component_ean) DO NOTHING;")
            comp_count += 1
    print()
    print(f"-- {comp_count} composition rows total")
    print()
    print("COMMIT;")
    print(f"-- Summary: {len(COMPONENTS)} components, {len(products)} composites, {comp_count} composition links", file=sys.stderr)


if __name__ == "__main__":
    main()
